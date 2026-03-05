#!/usr/bin/env python3
"""Upload ММ test data from Excel to Supabase."""

import sys
import json
import time
import openpyxl
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed

SUPABASE_URL = "https://zcciroutarcpkwpnynyh.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InpjY2lyb3V0YXJjcGt3cG55bnloIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzI3MjIzMTAsImV4cCI6MjA4ODI5ODMxMH0.LFnJ8WoxlNhZ06MBQm-1mmJK4mtkBLZAPd4UoPtGrkE"
HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=minimal",
}

EXCEL_PATH = "Demo data/ММ_данные_для_теста (1).xlsx"
BATCH_SIZE = 500
MAX_WORKERS = 1
MAX_RETRIES = 5


def safe_num(val):
    if val is None or str(val).strip() == "":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def safe_int(val):
    if val is None or str(val).strip() == "":
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def safe_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def parse_date(val):
    if val is None:
        return None
    s = str(val).strip()
    if " " in s:
        s = s.split(" ")[0]
    return s


def post_batch(table, rows, batch_num=0):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    for attempt in range(MAX_RETRIES):
        try:
            resp = requests.post(url, headers=HEADERS, json=rows, timeout=300)
            if resp.status_code in (200, 201, 204):
                return True
            print(f"  ERROR batch#{batch_num} status={resp.status_code}: {resp.text[:300]}", flush=True)
            if resp.status_code == 429:
                time.sleep(3 * (attempt + 1))
                continue
            return False
        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
            wait = 3 * (attempt + 1)
            print(f"  {type(e).__name__} batch#{batch_num} attempt {attempt+1}/{MAX_RETRIES}, wait {wait}s", flush=True)
            time.sleep(wait)
    return False


def upload_stores_and_products(wb):
    ws = wb["Исходник"]
    headers_list = None
    stores = {}
    products = {}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers_list = list(row)
            col = {h: idx for idx, h in enumerate(headers_list)}
            continue
        whs_id = str(row[col["WHS_ID"]])
        if whs_id not in stores:
            stores[whs_id] = {
                "whs_id": whs_id,
                "name": str(row[col["name_tt"]]),
                "code": str(row[col["code_tt"]]),
            }
        art_id = str(row[col["ART_ID"]])
        if art_id not in products:
            products[art_id] = {
                "art_id": art_id,
                "name_sku": str(row[col["name_sku"]]),
                "code_sku": str(row[col["code_sku"]]),
                "category_l0": safe_str(row[col["ART_GRP_LVL_0_NAME"]]),
                "category_l1": safe_str(row[col["ART_GRP_LVL_1_NAME"]]),
                "category_l2": safe_str(row[col["ART_GRP_LVL_2_NAME"]]),
                "category_l3": safe_str(row[col["ART_GRP_LVL_3_NAME"]]),
            }

    print(f"\n--- Uploading {len(stores)} stores ---", flush=True)
    store_list = list(stores.values())
    for i in range(0, len(store_list), BATCH_SIZE):
        batch = store_list[i : i + BATCH_SIZE]
        if not post_batch("stores", batch):
            return False
    print(f"  Done: {len(store_list)} stores uploaded", flush=True)

    print(f"\n--- Uploading {len(products)} products ---", flush=True)
    prod_list = list(products.values())
    for i in range(0, len(prod_list), BATCH_SIZE):
        batch = prod_list[i : i + BATCH_SIZE]
        if not post_batch("products", batch):
            return False
    print(f"  Done: {len(prod_list)} products uploaded", flush=True)
    return True


def _upload_batch(args):
    table, rows, batch_num = args
    return batch_num, post_batch(table, rows, batch_num)


def upload_source_data(wb):
    ws = wb["Исходник"]
    headers_list = None
    all_records = []

    print("\n--- Reading source_data (Исходник) ---", flush=True)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers_list = list(row)
            col = {h: idx for idx, h in enumerate(headers_list)}
            continue

        record = {
            "whs_id": str(row[col["WHS_ID"]]),
            "store_name": str(row[col["name_tt"]]),
            "store_code": str(row[col["code_tt"]]),
            "zone_name": str(row[col["zone_name"]]),
            "visit_date": parse_date(row[col["day_id"]]),
            "art_id": str(row[col["ART_ID"]]),
            "product_name": str(row[col["name_sku"]]),
            "product_code": str(row[col["code_sku"]]),
            "category_l0": safe_str(row[col["ART_GRP_LVL_0_NAME"]]),
            "category_l1": safe_str(row[col["ART_GRP_LVL_1_NAME"]]),
            "category_l2": safe_str(row[col["ART_GRP_LVL_2_NAME"]]),
            "category_l3": safe_str(row[col["ART_GRP_LVL_3_NAME"]]),
            "week_id": safe_str(row[col["WEEK_ID_2"]]),
            "is_test_week": safe_int(row[col["pr_test_week"]]),
            "sale_amount": safe_num(row[col["sale"]]),
            "sale_qty": safe_num(row[col["sale_qnty"]]),
            "stock_qty": safe_num(row[col["rest_qnty"]]),
            "days_count": safe_int(row[col["cnt_day"]]),
            "stock_qty_visit": safe_num(row[col["rest_qnty_vizit"]]),
            "in_target_assortment": safe_int(row[col["pr_ca"]]),
            "on_planogram": safe_int(row[col["pr_pg"]]),
            "sku_marker": safe_str(row[col["Признак SKU"]]),
            "face_width_planogram": safe_num(row[col["Фейс в шир по ПГ"]]),
            "face_width_actual": safe_num(row[col["Фейс в шир факт"]]),
        }
        all_records.append(record)

    print(f"  Read {len(all_records):,} records from Excel", flush=True)

    batches = []
    for i in range(0, len(all_records), BATCH_SIZE):
        batches.append(("source_data", all_records[i : i + BATCH_SIZE], i // BATCH_SIZE))

    print(f"  Uploading in {len(batches)} batches ({MAX_WORKERS} parallel)...", flush=True)

    errors = 0
    done = 0
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(_upload_batch, b): b[2] for b in batches}
        for future in as_completed(futures):
            batch_num, success = future.result()
            if not success:
                errors += 1
                if errors > 10:
                    print("  Too many errors, aborting", flush=True)
                    return False
            done += 1
            if done % 20 == 0:
                print(f"  Progress: {done}/{len(batches)} batches ({done * BATCH_SIZE:,} rows)...", flush=True)

    total = len(all_records)
    print(f"  Done: {total:,} rows uploaded ({errors} errors)", flush=True)
    return errors == 0


def upload_summary_data(wb):
    ws = wb["Свод"]
    all_records = []

    print("\n--- Reading summary_data (Свод) ---", flush=True)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 5:
            continue
        zone = safe_str(row[0])
        if not zone:
            continue

        record = {
            "zone_name": zone,
            "store_name": str(row[1]),
            "store_code": str(row[2]),
            "visit_date": parse_date(row[3]),
            "category_l0": safe_str(row[4]),
            "category_l1": safe_str(row[5]),
            "category_l2": safe_str(row[6]),
            "category_l3": safe_str(row[7]),
            "product_name": str(row[8]),
            "product_code": str(row[9]),
            "in_target_assortment": safe_int(row[10]),
            "on_planogram": safe_int(row[11]),
            "sku_marker": safe_str(row[12]),
            "face_width_planogram": safe_num(row[13]),
            "face_width_actual": safe_num(row[14]),
            "stock_qty_visit": safe_num(row[15]),
            "visit_week_sale_amount": safe_num(row[16]),
            "visit_week_sale_qty": safe_num(row[17]),
            "visit_week_avg_daily_stock": safe_num(row[18]),
            "other_weeks_sale_amount": safe_num(row[19]),
            "other_weeks_sale_qty": safe_num(row[20]),
            "other_weeks_avg_daily_stock": safe_num(row[21]),
        }
        all_records.append(record)

    print(f"  Read {len(all_records):,} records from Excel", flush=True)

    batches = []
    for i in range(0, len(all_records), BATCH_SIZE):
        batches.append(("summary_data", all_records[i : i + BATCH_SIZE], i // BATCH_SIZE))

    print(f"  Uploading in {len(batches)} batches ({MAX_WORKERS} parallel)...", flush=True)

    errors = 0
    done = 0
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(_upload_batch, b): b[2] for b in batches}
        for future in as_completed(futures):
            batch_num, success = future.result()
            if not success:
                errors += 1
                if errors > 10:
                    print("  Too many errors, aborting", flush=True)
                    return False
            done += 1
            if done % 10 == 0:
                print(f"  Progress: {done}/{len(batches)} batches...", flush=True)

    print(f"  Done: {len(all_records):,} rows uploaded ({errors} errors)", flush=True)
    return errors == 0


def main():
    target = sys.argv[1] if len(sys.argv) > 1 else "all"
    print(f"Loading Excel file: {EXCEL_PATH}", flush=True)
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}", flush=True)

    t0 = time.time()

    if target in ("all", "stores"):
        if not upload_stores_and_products(wb):
            print("FAILED: stores/products upload", flush=True)
            if target == "stores":
                return

    if target in ("all", "source"):
        if not upload_source_data(wb):
            print("FAILED: source_data upload", flush=True)

    if target in ("all", "summary"):
        if not upload_summary_data(wb):
            print("FAILED: summary_data upload", flush=True)

    elapsed = time.time() - t0
    print(f"\nTotal time: {elapsed:.1f}s", flush=True)
    wb.close()


if __name__ == "__main__":
    main()
